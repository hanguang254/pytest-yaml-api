# _*_ coding:utf-8 _*-
# @author：尹乐
import openpyxl

class PlayExcel():

    def __init__(self,filname,sheet):
        # 选择表单
        self.w = openpyxl.load_workbook(filname) #读取文件
        self.sh= self.w[sheet]



    #读所有内容,用列表加元祖显示
    def read_all(self):

        # 列
        max_column = self.sh.max_column
        #行
        max_row = self.sh.max_row

        l1 = []  #第二行开始
        for j in range(2, max_row + 1):
            l = []  #第三列开始
            for i in range(3, max_column + 1):
                table = self.sh.cell(row=j, column=i)
                l.append(table.value)
                # print(table.value,end=" ")

            l1.append(tuple(l))
            # print("")

        return l1



if __name__ == '__main__':
        data =   PlayExcel('../data/excel用例.xlsx','Sheet1').read_all()
        print(data)